"""Unit tests for the ``markdownize`` Prefect flow.

Covers the helpers, the manifest models, the single-file conversion task (via
``markitdown`` on a JSON input — local, no network), and smoke runs of
:func:`markdownize_flow`.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

import pytest

from genai_tk.workflow.flow_cache.manifest import ManifestCache
from genai_tk.workflow.markdownize.config import MarkdownizeProfile
from genai_tk.workflow.markdownize.flow import (
    MarkdownizeManifest,
    MarkdownizeManifestEntry,
    _chunked,
    _convert_file_task,
    markdownize_flow,
)
from genai_tk.workflow.markdownize.routing import (
    _is_markdownize_compatible,
    _output_paths,
    _prepare_files,
    _route_for,
)
from genai_tk.workflow.sources import ResolvedSourceFile

# ---------------------------------------------------------------------------
# _is_markdownize_compatible
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("suffix", "expected"),
    [
        (".pdf", True),
        (".docx", True),
        (".pptx", True),
        (".xlsx", True),
        (".xls", True),
        (".html", True),
        (".htm", True),
        (".csv", True),
        (".json", True),
        (".jpg", True),
        (".jpeg", True),
        (".png", True),
        (".gif", True),
        (".bmp", True),
        (".txt", False),
        (".md", True),
        (".markdown", True),
        (".mp3", False),
    ],
)
def test_is_markdownize_compatible(suffix: str, expected: bool) -> None:
    assert _is_markdownize_compatible(Path(f"file{suffix}")) is expected


# ---------------------------------------------------------------------------
# _chunked
# ---------------------------------------------------------------------------


def test_chunked_splits_into_batches() -> None:
    assert list(_chunked([1, 2, 3, 4, 5], 2)) == [[1, 2], [3, 4], [5]]


def test_chunked_zero_size_returns_single_batch() -> None:
    assert list(_chunked([1, 2, 3], 0)) == [[1, 2, 3]]


def test_chunked_empty() -> None:
    assert list(_chunked([], 3)) == []


# ---------------------------------------------------------------------------
# _prepare_files
# ---------------------------------------------------------------------------


def test_prepare_files_new_file_queued(tmp_path: Path) -> None:
    f = tmp_path / "a.pdf"
    f.write_bytes(b"%PDF fake")
    cache = ManifestCache()

    to_process, skipped = _prepare_files([ResolvedSourceFile(path=f, root=tmp_path)], cache, force=False)
    assert len(to_process) == 1
    assert skipped == 0
    assert to_process[0].path == f


def test_prepare_files_fresh_file_skipped(tmp_path: Path) -> None:
    f = tmp_path / "a.pdf"
    f.write_bytes(b"%PDF fake")
    cache = ManifestCache()
    from genai_tk.utils.hashing import buffer_digest

    cache.record_success(key=str(f), fingerprint=buffer_digest(f.read_bytes()), outputs={"output_path": "a_pdf.md"})

    to_process, skipped = _prepare_files([ResolvedSourceFile(path=f, root=tmp_path)], cache, force=False)
    assert to_process == []
    assert skipped == 1


def test_prepare_files_force_reprocesses_fresh(tmp_path: Path) -> None:
    f = tmp_path / "a.pdf"
    f.write_bytes(b"%PDF fake")
    cache = ManifestCache()
    from genai_tk.utils.hashing import buffer_digest

    cache.record_success(key=str(f), fingerprint=buffer_digest(f.read_bytes()), outputs={})

    to_process, skipped = _prepare_files([ResolvedSourceFile(path=f, root=tmp_path)], cache, force=True)
    assert len(to_process) == 1
    assert skipped == 0


def test_prepare_files_skips_unreadable(tmp_path: Path) -> None:
    cache = ManifestCache()
    to_process, skipped = _prepare_files(
        [ResolvedSourceFile(path=tmp_path / "missing.pdf", root=tmp_path)], cache, force=False
    )
    assert to_process == []
    assert skipped == 0


# ---------------------------------------------------------------------------
# MarkdownizeManifest — serialisation
# ---------------------------------------------------------------------------


def test_markdownize_manifest_round_trip() -> None:
    entry = MarkdownizeManifestEntry(
        source_hash="abc",
        output_path="out/a_pdf.md",
        processed_at=datetime(2024, 1, 1, tzinfo=timezone.utc),
    )
    manifest = MarkdownizeManifest(entries={"src/a.pdf": entry})
    recovered = MarkdownizeManifest.model_validate_json(manifest.model_dump_json())
    assert recovered.entries["src/a.pdf"].source_hash == "abc"
    assert recovered.entries["src/a.pdf"].output_path == "out/a_pdf.md"


def test_markdownize_manifest_empty_serialises() -> None:
    manifest = MarkdownizeManifest()
    data = json.loads(manifest.model_dump_json())
    assert data["entries"] == {}


# ---------------------------------------------------------------------------
# _route_for
# ---------------------------------------------------------------------------


def test_route_for_uses_profile_per_family() -> None:
    profile = MarkdownizeProfile(
        ppt_converter="via_pdf", doc_converter="markitdown", excel_converter="md_parser", pdf_converter="mistral"
    )
    assert _route_for(Path("deck.pptx"), profile) == "via_pdf"
    assert _route_for(Path("memo.docx"), profile) == "markitdown"
    assert _route_for(Path("sheet.xlsx"), profile) == "md_parser"
    assert _route_for(Path("scan.pdf"), profile) == "pdf"
    assert _route_for(Path("photo.png"), profile) == "markitdown"


# ---------------------------------------------------------------------------
# _convert_file_task (markitdown / md_parser — local conversion)
# ---------------------------------------------------------------------------


@pytest.mark.fake_models
def test_convert_file_task_converts_json_to_markdown(tmp_path: Path) -> None:
    src = tmp_path / "data.json"
    src.write_text('{"name": "Ada", "age": 30}', encoding="utf-8")
    out = tmp_path / "out"

    rel_out, out_abs = _output_paths(src, tmp_path, out)
    source, relative_output = _convert_file_task.fn(
        str(src), str(src), "markitdown", str(out_abs), str(rel_out), "markitdown"
    )

    assert source == str(src)
    assert relative_output == "data_json.md"
    written = out / "data_json.md"
    assert written.exists()
    assert "Ada" in written.read_text(encoding="utf-8")


@pytest.mark.fake_models
def test_convert_file_task_preserves_subdir_structure(tmp_path: Path) -> None:
    (tmp_path / "docs").mkdir()
    src = tmp_path / "docs" / "data.csv"
    src.write_text("name,age\nAda,30\n", encoding="utf-8")
    out = tmp_path / "out"

    rel_out, out_abs = _output_paths(src, tmp_path, out)
    source, relative_output = _convert_file_task.fn(
        str(src), str(src), "markitdown", str(out_abs), str(rel_out), "markitdown"
    )

    assert source == str(src)
    assert relative_output == str(Path("docs") / "data_csv.md")
    assert (out / "docs" / "data_csv.md").exists()


# ---------------------------------------------------------------------------
# Excel conversion (md-spreadsheet-parser)
# ---------------------------------------------------------------------------


def _write_messy_workbook(path: Path) -> None:
    """Write an .xlsx with a merged title row, a blank spacer row, and empty cells."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Sales Report"
    ws.merge_cells("A1:C1")
    ws["A3"] = "Region"
    ws["B3"] = "Jan"
    ws["C3"] = "Feb"
    ws["A4"] = "North"
    ws["B4"] = 100
    ws["A5"] = "South"
    ws["C5"] = 80
    wb.save(path)


def test_excel_to_markdown_md_parser_no_nan_and_title_rescued(tmp_path: Path) -> None:
    from genai_tk.workflow.markdownize.excel import _excel_to_markdown_md_parser

    src = tmp_path / "sample.xlsx"
    _write_messy_workbook(src)

    content = _excel_to_markdown_md_parser(src)

    assert "NaN" not in content
    assert "### Sales Report" in content
    assert "## Sales" in content
    assert "| Region | Jan | Feb |" in content


@pytest.mark.fake_models
def test_convert_file_task_routes_xlsx_through_md_parser(tmp_path: Path) -> None:
    src = tmp_path / "sample.xlsx"
    _write_messy_workbook(src)
    out = tmp_path / "out"

    rel_out, out_abs = _output_paths(src, tmp_path, out)
    source, relative_output = _convert_file_task.fn(
        str(src), str(src), "md_parser", str(out_abs), str(rel_out), "markitdown"
    )

    assert source == str(src)
    written = out / relative_output
    content = written.read_text(encoding="utf-8")
    assert "NaN" not in content
    assert "### Sales Report" in content


@pytest.mark.fake_models
def test_convert_file_task_xlsx_falls_back_to_markitdown_on_parser_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import genai_tk.workflow.markdownize.converters as converters_module

    def _boom(path: Path) -> str:
        raise ValueError("boom")

    monkeypatch.setattr(converters_module, "_excel_to_markdown_md_parser", _boom)

    src = tmp_path / "sample.xlsx"
    _write_messy_workbook(src)
    out = tmp_path / "out"

    rel_out, out_abs = _output_paths(src, tmp_path, out)
    source, relative_output = _convert_file_task.fn(
        str(src), str(src), "md_parser", str(out_abs), str(rel_out), "markitdown"
    )

    assert source == str(src)
    assert (out / relative_output).exists()


# ---------------------------------------------------------------------------
# markdownize_flow — smoke runs
# ---------------------------------------------------------------------------


@pytest.mark.fake_models
def test_markdownize_flow_no_files_returns_empty_manifest(tmp_path: Path) -> None:
    src = tmp_path / "src"
    src.mkdir()
    (src / "notes.txt").write_text("not supported", encoding="utf-8")
    out = tmp_path / "out"

    manifest = markdownize_flow(sources=str(src), md_output_dir=str(out))
    assert isinstance(manifest, MarkdownizeManifest)
    assert manifest.entries == {}


@pytest.mark.fake_models
def test_markdownize_flow_converts_and_writes_manifest(tmp_path: Path) -> None:
    src = tmp_path / "src"
    src.mkdir()
    (src / "data.json").write_text('{"name": "Ada", "age": 30}', encoding="utf-8")
    out = tmp_path / "out"

    manifest = markdownize_flow(sources=str(src), md_output_dir=str(out), batch_size=1)

    assert isinstance(manifest, MarkdownizeManifest)
    assert len(manifest.entries) == 1
    assert (out / ".cache" / "manifest.json").exists()
    assert (out / "data_json.md").exists()
    assert "Ada" in (out / "data_json.md").read_text(encoding="utf-8")


@pytest.mark.fake_models
def test_markdownize_flow_copies_existing_markdown(tmp_path: Path) -> None:
    src = tmp_path / "src"
    src.mkdir()
    (src / "readme.md").write_text("# Hello\n", encoding="utf-8")
    out = tmp_path / "out"

    manifest = markdownize_flow(sources=str(src), md_output_dir=str(out), batch_size=1)

    assert len(manifest.entries) == 1
    assert (out / "readme.md").read_text(encoding="utf-8") == "# Hello\n"


@pytest.mark.fake_models
def test_markdownize_flow_all_cached_returns_without_reprocessing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    src = tmp_path / "src"
    src.mkdir()
    f = src / "data.json"
    f.write_text('{"name": "Ada"}', encoding="utf-8")
    out = tmp_path / "out"
    cache_dir = out / ".cache"
    cache_dir.mkdir(parents=True)

    from genai_tk.utils.hashing import buffer_digest
    from genai_tk.workflow.markdownize.config import get_markdownize_profile

    code_version = get_markdownize_profile("default").fingerprint()
    cache = ManifestCache.load(cache_dir / "manifest.json")
    cache.record_success(
        key=str(f),
        fingerprint=buffer_digest(f.read_bytes()),
        code_version=code_version,
        outputs={"output_path": "data_json.md"},
    )
    cache.save(cache_dir / "manifest.json")

    # If the flow reprocesses, markitdown runs — prevent it to prove caching works.
    def _boom(*args, **kwargs):  # noqa: ANN002
        raise AssertionError("should not reconvert a cached file")

    monkeypatch.setattr("genai_tk.workflow.markdownize.converters._markitdown_text", _boom, raising=False)

    manifest = markdownize_flow(sources=str(src), md_output_dir=str(out))
    assert str(f) in manifest.entries
