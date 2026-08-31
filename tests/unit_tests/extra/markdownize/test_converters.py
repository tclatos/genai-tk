"""Unit tests for document converter implementations in genai_tk.extra.markdownize."""

from __future__ import annotations

import sys
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from genai_tk.extra.markdownize.anydoc_converter import AnyDocConverter
from genai_tk.extra.markdownize.edgeparse_converter import EdgeParseConverter
from genai_tk.extra.markdownize.excel_converter import MessyExcelConverter
from genai_tk.extra.markdownize.lighton_ocr_converter import LightOnOCRConverter
from genai_tk.extra.markdownize.llm_converter import LLMConverter
from genai_tk.extra.markdownize.markitdown_converter import MarkItDownConverter
from genai_tk.extra.markdownize.mistral_ocr_converter import MistralOCRConverter


@pytest.mark.asyncio
async def test_markitdown_converter(tmp_path: Path) -> None:
    test_file = tmp_path / "test.json"
    test_file.write_text('{"project": "genai-tk", "version": "1.0"}', encoding="utf-8")

    converter = MarkItDownConverter()
    assert ".json" in converter.supported_extensions()
    assert ".pdf" in converter.supported_extensions()

    md = await converter.convert(test_file)
    assert "genai-tk" in md


@pytest.mark.asyncio
async def test_messy_excel_converter(tmp_path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Q1 Summary"
    ws.merge_cells("A1:B1")
    ws["A2"] = "Metric"
    ws["B2"] = "Value"
    ws["A3"] = "Revenue"
    ws["B3"] = 1000

    test_file = tmp_path / "sheet.xlsx"
    wb.save(test_file)

    converter = MessyExcelConverter()
    assert ".xlsx" in converter.supported_extensions()

    md = await converter.convert(test_file)
    assert "Q1 Summary" in md
    assert "Revenue" in md


@pytest.mark.asyncio
async def test_edgeparse_converter(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    test_file = tmp_path / "doc.pdf"
    test_file.write_bytes(b"%PDF-1.4 fake")

    fake_edgeparse = MagicMock()
    fake_edgeparse.convert.return_value = "# Parsed by EdgeParse"
    monkeypatch.setattr("sys.modules", {**sys.modules, "edgeparse": fake_edgeparse})  # type: ignore

    converter = EdgeParseConverter()
    assert ".pdf" in converter.supported_extensions()
    with patch.dict("sys.modules", {"edgeparse": fake_edgeparse}):
        md = await converter.convert(test_file)
        assert md == "# Parsed by EdgeParse"


@pytest.mark.asyncio
async def test_edgeparse_converter_missing_package(tmp_path: Path) -> None:
    test_file = tmp_path / "doc.pdf"
    test_file.write_bytes(b"%PDF-1.4 fake")

    converter = EdgeParseConverter()
    with patch.dict("sys.modules", {"edgeparse": None}):
        with pytest.raises(ImportError, match="edgeparse is required"):
            await converter.convert(test_file)


@pytest.mark.asyncio
async def test_mistral_ocr_converter_single(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    test_file = tmp_path / "sample.pdf"
    test_file.write_bytes(b"%PDF-1.4 sample")

    fake_page = MagicMock()
    fake_page.index = 0
    fake_page.markdown = "# Page One Content"
    fake_response = MagicMock(pages=[fake_page])

    fake_client = MagicMock()
    fake_client.ocr.process.return_value = fake_response

    converter = MistralOCRConverter(api_key="fake-key")
    monkeypatch.setattr(converter, "_get_client", lambda: fake_client)

    md = await converter.convert(test_file)
    assert "## Page 1" in md
    assert "# Page One Content" in md


@pytest.mark.asyncio
async def test_mistral_ocr_converter_batch(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    f1 = tmp_path / "1.pdf"
    f2 = tmp_path / "2.pdf"
    f1.write_bytes(b"%PDF-1.4 1")
    f2.write_bytes(b"%PDF-1.4 2")

    converter = MistralOCRConverter(api_key="fake-key", use_batch_api=True)
    fake_client = MagicMock()
    monkeypatch.setattr(converter, "_get_client", lambda: fake_client)

    async def _mock_submit(client, requests, files):
        return {str(f1): "## Page 1\n\nContent 1", str(f2): "## Page 1\n\nContent 2"}

    monkeypatch.setattr(converter, "_submit_and_poll_batch", _mock_submit)

    results = await converter.batch_convert([f1, f2])
    assert str(f1) in results
    assert str(f2) in results
    assert results[str(f1)] == "## Page 1\n\nContent 1"


@pytest.mark.asyncio
async def test_lighton_ocr_converter_sync(tmp_path: Path) -> None:
    test_file = tmp_path / "invoice.pdf"
    test_file.write_bytes(b"%PDF-1.4 invoice")

    converter = LightOnOCRConverter(api_key="fake-lighton-key", async_mode=False)
    assert ".pdf" in converter.supported_extensions()
    assert ".docx" in converter.supported_extensions()

    mock_resp = MagicMock()
    mock_resp.status_code = 200
    mock_resp.json.return_value = {
        "id": "parse_123",
        "status": "completed",
        "result": {
            "pages": [
                {"index": 1, "markdown": "# Invoice Details\nTotal: $500"},
            ]
        },
    }

    mock_http_client = AsyncMock()
    mock_http_client.post.return_value = mock_resp
    mock_http_client.__aenter__.return_value = mock_http_client

    with patch("httpx.AsyncClient", return_value=mock_http_client):
        md = await converter.convert(test_file)
        assert "# Invoice Details" in md
        assert "Total: $500" in md


@pytest.mark.asyncio
async def test_lighton_ocr_converter_async_polling(tmp_path: Path) -> None:
    test_file = tmp_path / "large.pdf"
    test_file.write_bytes(b"%PDF-1.4 large")

    converter = LightOnOCRConverter(api_key="fake-key", async_mode=True, poll_interval_seconds=0.01)

    accepted_resp = MagicMock()
    accepted_resp.status_code = 202
    accepted_resp.json.return_value = {"id": "parse_job_99"}

    poll_completed_resp = MagicMock()
    poll_completed_resp.status_code = 200
    poll_completed_resp.json.return_value = {
        "status": "completed",
        "result": {
            "pages": [
                {"index": 1, "markdown": "# Page 1"},
                {"index": 2, "markdown": "# Page 2"},
            ]
        },
    }

    mock_http_client = AsyncMock()
    mock_http_client.post.return_value = accepted_resp
    mock_http_client.get.return_value = poll_completed_resp
    mock_http_client.__aenter__.return_value = mock_http_client

    with patch("httpx.AsyncClient", return_value=mock_http_client):
        md = await converter.convert(test_file)
        assert "## Page 1" in md
        assert "## Page 2" in md


@pytest.mark.asyncio
async def test_anydoc_converter_mocked(tmp_path: Path) -> None:
    test_file = tmp_path / "report.docx"
    test_file.write_bytes(b"PK fake docx")

    converter = AnyDocConverter()
    assert ".docx" in converter.supported_extensions()
    assert ".pdf" in converter.supported_extensions()

    fake_anydoc = MagicMock()
    fake_anydoc.to_markdown.return_value = "# Converted by AnyDoc"

    with patch.dict("sys.modules", {"anydoc": fake_anydoc}):
        md = await converter.convert(test_file)
        assert md == "# Converted by AnyDoc"


@pytest.mark.asyncio
async def test_anydoc_converter_missing_package(tmp_path: Path) -> None:
    test_file = tmp_path / "report.docx"
    test_file.write_bytes(b"PK fake docx")

    converter = AnyDocConverter()
    with patch.dict("sys.modules", {"anydoc": None}):
        with pytest.raises(ImportError, match="firecrawl-anydoc is required"):
            await converter.convert(test_file)


@pytest.mark.asyncio
async def test_llm_converter_text(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    test_file = tmp_path / "sample.txt"
    test_file.write_text("Title: Meeting Notes\nAttendees: Alice, Bob", encoding="utf-8")

    mock_llm = MagicMock()
    mock_response = MagicMock(content="# Meeting Notes\n\n- Alice\n- Bob")
    mock_llm.ainvoke = AsyncMock(return_value=mock_response)

    converter = LLMConverter(llm="fake_llm")
    monkeypatch.setattr(converter, "_get_model", lambda: mock_llm)

    md = await converter.convert(test_file)
    assert "# Meeting Notes" in md
    assert "- Alice" in md


@pytest.mark.asyncio
async def test_llm_converter_binary(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    test_file = tmp_path / "image.png"
    test_file.write_bytes(b"\x89PNG\r\n\x1a\nfake")

    mock_llm = MagicMock()
    mock_response = MagicMock(content="# Diagram Description\n\nFlowchart text")
    mock_llm.ainvoke = AsyncMock(return_value=mock_response)

    converter = LLMConverter(llm="fake_vision")
    monkeypatch.setattr(converter, "_get_model", lambda: mock_llm)

    md = await converter.convert(test_file)
    assert "# Diagram Description" in md


@pytest.mark.asyncio
async def test_llm_converter_batch(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    f1 = tmp_path / "1.txt"
    f2 = tmp_path / "2.txt"
    f1.write_text("doc 1", encoding="utf-8")
    f2.write_text("doc 2", encoding="utf-8")

    mock_llm = MagicMock()
    mock_llm.ainvoke = AsyncMock(
        side_effect=[
            MagicMock(content="# Output 1"),
            MagicMock(content="# Output 2"),
        ]
    )

    converter = LLMConverter(llm="fake_llm", max_concurrency=2)
    monkeypatch.setattr(converter, "_get_model", lambda: mock_llm)

    results = await converter.batch_convert([f1, f2])
    assert results[str(f1)] == "# Output 1"
    assert results[str(f2)] == "# Output 2"
