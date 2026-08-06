"""Deterministic, dependency-free ``.xlsx``/``.xls`` to Markdown conversion for messy real-world sheets.

Unlike a naive "dump every cell into one big table" converter, this module
first segments each worksheet into independent regions and formats each region
according to what it actually looks like. Real-world spreadsheets routinely mix
the following patterns within a single sheet, all of which are handled here:

- **Merged title/banner cells** — a single cell merged across a row (e.g. a
  report title, or a section banner above a table) is recovered as a heading
  or promoted title rather than one cell followed by empty ones.
- **Grouped / multi-row headers** — column groups such as a ``Q1`` cell
  merged over ``Jan``/``Feb`` sub-columns are folded into one header row
  (``Q1 Jan``, ``Q1 Feb``) instead of leaking the sub-header as a data row.
- **Multiple tables stacked vertically** in one sheet, separated by blank rows
  (e.g. a summary table followed by a detail table below it).
- **Multiple tables placed side-by-side** in one sheet, separated by blank
  spacer columns (a common layout for compact reference/lookup sheets).
- **Free-text blocks** — instructions, legends, or key/value note lists
  (narrow, 1-2 filled cells per row) rendered as prose/bullets instead of
  being forced into a table.
- **Dates, times, percentages, booleans, and float-formatted numbers** —
  rendered using the cell's actual value/format rather than raw ``repr``
  noise (``0.15`` with a ``%`` format renders as ``15%``, not ``0.15``).
- **Cell values containing newlines or ``|``** — escaped so the emitted
  Markdown table stays well-formed.
- **Hidden worksheets** — skipped, since they're rarely meant for a reader.
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any


def _format_datetime(value: datetime.date | datetime.time) -> str:
    """Render a date/time value, collapsing midnight datetimes to a plain date."""
    if isinstance(value, datetime.datetime):
        if value.hour or value.minute or value.second or value.microsecond:
            return value.isoformat(sep=" ")
        return value.date().isoformat()
    return value.isoformat()


def _format_percent(value: float, number_format: str) -> str:
    """Render a numeric value as a percentage, honouring the format's decimal places."""
    decimals = 0
    if "." in number_format:
        fraction = number_format.split("%")[0].split(".")[-1]
        decimals = sum(1 for ch in fraction if ch in "0#")
    return f"{value * 100:.{decimals}f}%"


def _format_number(value: float) -> str:
    """Render a number, dropping float noise and integer-like trailing decimals."""
    if value != value or value in (float("inf"), float("-inf")):  # NaN / inf
        return ""
    rounded = round(value, 10)
    if rounded == int(rounded):
        return str(int(rounded))
    return f"{rounded}"


def _format_cell(cell: Any) -> str:
    """Stringify an openpyxl cell, applying date/percent/number formatting."""
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return _format_datetime(value)
    if isinstance(value, (int, float)):
        number_format = (getattr(cell, "number_format", "") or "").lower()
        if "%" in number_format:
            return _format_percent(float(value), number_format)
        return _format_number(float(value))
    return str(value).strip()


def _build_grid(worksheet: Any) -> list[list[str]]:
    """Read a worksheet into a string grid, filling merged ranges with their anchor value."""
    merged_anchor: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in worksheet.merged_cells.ranges:
        anchor = (rng.min_row, rng.min_col)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != anchor:
                    merged_anchor[(r, c)] = anchor

    grid: list[list[str]] = []
    for row in worksheet.iter_rows():
        cells: list[str] = []
        for cell in row:
            anchor = merged_anchor.get((cell.row, cell.column))
            source = worksheet.cell(row=anchor[0], column=anchor[1]) if anchor else cell
            cells.append(_format_cell(source))
        grid.append(cells)
    return grid


def _is_row_empty(row: list[str]) -> bool:
    """Return True if all cells in the row are empty or whitespace."""
    return not any(cell.strip() for cell in row)


def _split_sheet_into_vertical_blocks(grid: list[list[str]]) -> list[list[list[str]]]:
    """Split a sheet grid into contiguous blocks separated by empty rows."""
    blocks: list[list[list[str]]] = []
    current_block: list[list[str]] = []

    for row in grid:
        if _is_row_empty(row):
            if current_block:
                blocks.append(current_block)
                current_block = []
        else:
            current_block.append(row)

    if current_block:
        blocks.append(current_block)

    return blocks


def _split_block_into_horizontal_blocks(block: list[list[str]]) -> list[list[list[str]]]:
    """Split a block into side-by-side sub-blocks separated by fully-empty columns."""
    if not block:
        return []

    width = max(len(row) for row in block)
    padded = [row + [""] * (width - len(row)) for row in block]
    col_has_data = [any(padded[r][c].strip() for r in range(len(padded))) for c in range(width)]

    subblocks: list[list[list[str]]] = []
    c = 0
    while c < width:
        if not col_has_data[c]:
            c += 1
            continue
        start = c
        while c < width and col_has_data[c]:
            c += 1
        subblocks.append([row[start:c] for row in padded])

    return subblocks or [padded]


def _clean_block_columns(block: list[list[str]]) -> list[list[str]]:
    """Trim empty leading/trailing/spacer columns and pad ragged rows to equal width."""
    if not block:
        return []

    max_width = max(len(row) for row in block)
    padded_block = [row + [""] * (max_width - len(row)) for row in block]

    keep_cols = [
        col_idx
        for col_idx in range(max_width)
        if any(padded_block[row_idx][col_idx].strip() for row_idx in range(len(padded_block)))
    ]

    if not keep_cols:
        return []

    return [[row[col_idx] for col_idx in keep_cols] for row in padded_block]


def _dedup_consecutive(row: list[str]) -> list[str]:
    """Return non-empty cells with consecutive duplicates collapsed (undoes merge fill)."""
    result: list[str] = []
    previous: str | None = None
    for cell in row:
        text = cell.strip()
        if not text:
            previous = None
            continue
        if text != previous:
            result.append(text)
        previous = text
    return result


def _has_horizontal_span(row: list[str]) -> bool:
    """Return True if the row has two identical adjacent cells (a horizontal merge)."""
    previous: str | None = None
    for cell in row:
        text = cell.strip()
        if text and text == previous:
            return True
        previous = text or previous
    return False


def _collapse_header_rows(table_grid: list[list[str]]) -> list[list[str]]:
    """Fold leading grouped-header rows (from horizontal merges) into one header row.

    Consecutive leading rows carrying horizontal spans plus the first plain row
    are joined column-wise (e.g. ``Q1`` over ``Jan`` becomes ``Q1 Jan``).
    """
    span_rows = 0
    for row in table_grid:
        if _has_horizontal_span(row):
            span_rows += 1
        else:
            break

    if span_rows == 0 or span_rows >= len(table_grid):
        return table_grid

    header_rows = table_grid[: span_rows + 1]
    body = table_grid[span_rows + 1 :]
    width = len(table_grid[0])

    merged_header: list[str] = []
    for col in range(width):
        parts: list[str] = []
        previous: str | None = None
        for hr in header_rows:
            text = hr[col].strip() if col < len(hr) else ""
            if text and text != previous:
                parts.append(text)
            previous = text or previous
        merged_header.append(" ".join(parts))

    return [merged_header, *body]


def _escape_md_cell(text: str) -> str:
    """Make a cell value safe inside a Markdown table (newlines and pipes)."""
    return _to_br(text).replace("|", "\\|")


def _to_br(text: str) -> str:
    """Replace embedded line breaks with ``<br>`` so multi-line cell text renders on one block."""
    return text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")


def _classify_and_format_block(block: list[list[str]]) -> str:
    """Determine whether a block is text/heading or a table, and convert to Markdown."""
    block = _clean_block_columns(block)
    if not block:
        return ""

    num_rows = len(block)
    # Logical cells per row collapse merge-fill duplicates so layout heuristics stay honest.
    logical = [_dedup_consecutive(row) for row in block]

    # --- Case 1: Single line -> bold lead-in or paragraph (no sub-heading: one section per sheet) ---
    if num_rows == 1 and len(logical[0]) == 1:
        text = _to_br(logical[0][0])
        if len(text) < 80 and not text.endswith("."):
            return f"**{text}**\n"
        return f"{text}\n"

    # --- Case 2: Narrow multi-row text block (instructions, legend, notes, key/value) ---
    logical_counts = [len(cells) for cells in logical]
    max_logical = max(logical_counts)
    avg_logical = sum(logical_counts) / num_rows

    if avg_logical <= 1.5 and max_logical <= 2:
        lines: list[str] = []
        for cells in logical:
            text = _to_br(" ".join(cells))
            if not text:
                continue
            if "=" in text or text.startswith("-"):
                lines.append(f"- {text}")
            else:
                lines.append(text)
        return "\n".join(lines) + "\n"

    # --- Case 3: Table Block (optionally preceded by a single-cell title row) ---
    title: str | None = None
    table_grid = block
    if num_rows > 1 and len(logical[0]) == 1 and len(logical[1]) > 1:
        title = logical[0][0]
        table_grid = block[1:]

    table_grid = _collapse_header_rows(table_grid)

    md_lines: list[str] = []
    if title:
        md_lines.append(f"**{_to_br(title)}**\n")

    headers = table_grid[0]
    headers_clean = [_escape_md_cell(h) if h.strip() else f"Col {i + 1}" for i, h in enumerate(headers)]

    md_lines.append("| " + " | ".join(headers_clean) + " |")
    md_lines.append("| " + " | ".join(["---"] * len(headers_clean)) + " |")

    for row in table_grid[1:]:
        md_lines.append("| " + " | ".join(_escape_md_cell(c) for c in row) + " |")

    return "\n".join(md_lines) + "\n"


def excel_to_markdown_messy_xls(path: Path) -> str:
    """Convert an Excel spreadsheet to Markdown, tolerating messy real-world layouts.

    Handles merged cells, grouped headers, headings, introductory text, legend
    blocks, spacer columns, and multiple stacked or side-by-side tables per
    worksheet. Hidden worksheets are skipped.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    md_parts: list[str] = []

    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            continue

        grid = _build_grid(worksheet)
        if not any(any(c.strip() for c in row) for row in grid):
            continue

        md_parts.append(f"## Sheet: {worksheet.title}\n")

        for block in _split_sheet_into_vertical_blocks(grid):
            for subblock in _split_block_into_horizontal_blocks(block):
                formatted_block = _classify_and_format_block(subblock)
                if formatted_block.strip():
                    md_parts.append(formatted_block)

    return "\n\n".join(md_parts)
